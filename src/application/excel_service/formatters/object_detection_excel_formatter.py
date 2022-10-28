import os
from typing import List
import numpy as np
from openpyxl.styles import Alignment, Font, Border, Side
from pandas import DataFrame
from sklearn.metrics import confusion_matrix
from statistics import mean
import pandas as pd
import openpyxl
import xlsxwriter
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook

from application.excel_service.models.linkage_definition import LinkageDefinition
from domain.contracts.abstract_excel_formatter import AbstractExcelFormatter
from domain.models.evaluation_job_parameters import EvaluationJobParameters
from domain.models.linkage_type import LinkageType


class ObjectDetectionExcelFormatter(AbstractExcelFormatter):
    def __init__(self):
        self.thick_border = Border(left=Side(style='thick'),
                                   right=Side(style='thick'),
                                   top=Side(style='thick'),
                                   bottom=Side(style='thin'))

        self.top_thick_border = Border(top=Side(style='thick'))
        self.bottom_thick_border = Border(bottom=Side(style='thick'))
        self.thin_border = Border(left=Side(style='thin'),
                                  right=Side(style='thin'),
                                  top=Side(style='thin'),
                                  bottom=Side(style='thin'))

        self.thick_bottom = Border(left=Side(style='thin'),
                                   right=Side(style='thin'),
                                   top=Side(style='thin'),
                                   bottom=Side(style='thick'))

    def get_object_detection_detailed_df(self, linkages_df: DataFrame, prediction_df: DataFrame, gt_df: DataFrame) -> DataFrame:
        all_df: DataFrame = linkages_df.copy()
        all_gt_index: List[int] = all_df.gt_index.tolist()
        all_pred_index: List[int] = all_df.prediction_index.tolist()

        gt_bbox = gt_df.loc[gt_df.index.isin(all_gt_index), ['left', 'right', 'top', 'bottom']].rename(
            columns={'left': 'gt_left', 'right': 'gt_right', 'top': 'gt_top', "bottom": "gt_bottom"})
        pred_bbox = prediction_df.loc[prediction_df.index.isin(all_pred_index), ['left', 'right', 'top', 'bottom']].rename(
            columns={'left': 'pred_left', 'right': 'pred_right', 'top': 'pred_top', "bottom": "pred_bottom"})

        df_all_cols = pd.concat([all_df, gt_bbox, pred_bbox], axis=1)
        df_all_cols.drop(['id'], axis=1, inplace=True)
        df_all_cols.drop(['gt_index', 'prediction_index'], axis=1, inplace=True)

        df_all_cols = df_all_cols[
            ['image_path', 'gt_label', 'prediction_label', 'linkage_type', 'iou', 'ecludien_distance', 'confidence', 'gt_centroid', 'prediction_centroid',
             'gt_left', 'gt_right', 'gt_top', 'gt_bottom', 'pred_left', 'pred_right', 'pred_top', 'pred_bottom']]
        df_all_cols = df_all_cols.dropna()
        df_all_cols.image_path = df_all_cols.image_path.apply(lambda x: os.path.basename(x))
        df_all_cols.linkage_type = df_all_cols.linkage_type.apply(lambda x: LinkageDefinition[x].value)

        df_all_cols[['iou', 'ecludien_distance', 'confidence']] = df_all_cols[['iou', 'ecludien_distance', 'confidence']].astype(float).round(2)
        df_all_cols.fillna(value='--', inplace=True)
        df_all_cols.sort_values('image_path', inplace=True)
        df_all_cols.reset_index(drop=True, inplace=True)

        return df_all_cols

    def add_general_header_info(self, worksheet, linkages_df: DataFrame, evaluation_job_parameter: EvaluationJobParameters):
        wrap_text = Alignment(wrapText=True)
        worksheet['A1'].alignment = wrap_text
        worksheet['A1'].font = Font(b=True, name="Tahoma")
        worksheet['A1'] = "Evaluation Job Per Image Report"
        worksheet['A2'] = 'Evaluation Job Name : ' + str(evaluation_job_parameter.job_name)
        worksheet['A3'] = 'Inference URL : ' + str(evaluation_job_parameter.url)
        worksheet['A4'] = 'Model Name : ' + str(evaluation_job_parameter.model_name)
        worksheet['A5'] = 'Total Image Tested : ' + str(len(linkages_df.image_path.unique()))
        for i in range(1, 6):
            worksheet.cell(row=i, column=1).border = self.thin_border

        return worksheet

    # def _add_header_image(self, worksheet):
    #     worksheet.row_dimensions[1].height = 80
    #     img = openpyxl.drawing.image.Image('./BMW_Innovation_Lab_Logo.png')
    #     img.anchor = 'A1'
    #     img.width = 100
    #     img.height = 100
    #     worksheet.add_image(img)
    #     worksheet["A1"].alignment = Alignment(horizontal="center", wrapText=True)
    #
    #     return worksheet

    def _add_header_info(self, worksheet, evaluation_job_name, evaluation_url, evaluation_model_name):
        wrap_text = Alignment(wrapText=True)
        worksheet['A2'].alignment = wrap_text
        worksheet['A2'].font = Font(b=True, name="Tahoma")
        worksheet['A2'] = "Evaluation Job Per Image Report"
        worksheet['A3'] = 'Evaluation Job Name : ' + str(evaluation_job_name)
        worksheet['A4'] = 'Inference URL : ' + str(evaluation_url)
        worksheet['A5'] = 'Model Name : ' + str(evaluation_model_name)

        return worksheet

    def get_multi_class_confusion_df(self, linkages_df: DataFrame) -> DataFrame:
        prediction_labels: List[str] = linkages_df.prediction_label.astype(str).tolist()
        gt_labels: List[str] = linkages_df.gt_label.astype(str).tolist()
        all_labels: List[str] = list(set(prediction_labels + gt_labels))

        # add nan to the end in the case of object_detection
        if "nan" in all_labels:
            all_labels.append(all_labels.pop(all_labels.index("nan")))
        all_labels.sort()
        confusion_matrix_output = confusion_matrix(gt_labels, prediction_labels, labels=all_labels).tolist()
        return pd.DataFrame(data=confusion_matrix_output, columns=all_labels, index=all_labels)

    def get_confusion_matrix_df(self, linkages_df: DataFrame) -> DataFrame:
        columns: List[str] = ['No Prediction', 'Prediction']
        indexes: List[str] = ['No True Label', 'True Label']
        labels: List[str] = [linkage_type.value for linkage_type in LinkageType]
        data = [(linkages_df.linkage_type.values == label).sum().tolist() for label in labels]
        data = np.asarray(data).reshape(2, 2)[::-1]
        data = np.fliplr(data)
        return pd.DataFrame(data=data, columns=columns, index=indexes)

    def get_statistics_df(self, linkages_df: DataFrame, predictions_df: DataFrame, gt_df: DataFrame, bad_iou) -> DataFrame:
        TP, TN, FP, FN = [(linkages_df.linkage_type.values == case.value).sum() for case in LinkageType]
        try:
            precision: float = round(TP / (TP + FP), 2)
            recall: float = round(TP / (TP + FN), 2)
            f_score: float = round(2 * precision * recall / (precision + recall), 2)
        except Exception:
            precision = recall = f_score = 0
        try:
            TNR: float = TN / float(TN + FP)
            TPR: float = TP / float(TP + FN)
        except Exception:
            TNR = TPR = 0
        try:
            accuracy = round((TP + TN) / float(TP + TN + FP + FN), 2)
            balanced_accuracy = round(float(TNR + TPR) / 2, 2)
        except Exception:
            accuracy = balanced_accuracy = 0
        try:
            avg_iou = round(mean(linkages_df.loc[linkages_df.iou.notnull(), 'iou'].tolist()), 2)
            avg_conf = round(mean(linkages_df.loc[linkages_df.confidence.notnull(), 'confidence'].tolist()), 2)
        except Exception:
            avg_conf = avg_iou = 0

        len_gt = len(gt_df.index)
        len_pred = len(predictions_df.index)

        indexes = ['Total number of predictions', 'Total number of Ground truths', 'Bad IoU Threshold (Wrong Label Case)',
                   'Accuracy', 'Balanced Accuracy', 'Precision', 'Recall', 'F-score', 'Average IoU',
                   'Average Confidence']
        data = [len_pred, len_gt, bad_iou, accuracy, balanced_accuracy, precision, recall, f_score, avg_iou, avg_conf]
        return pd.DataFrame(data=data, index=indexes, columns=['score'])

    def get_iou_df(self, linkages_df: DataFrame, evaluation_job_parameter: EvaluationJobParameters) -> DataFrame:
        good_iou: float = evaluation_job_parameter.iou_good_threshold
        avg_iou: float = evaluation_job_parameter.iou_average_threshold

        bad_iou_title: str = "Bad IoU: IoU<" + str(avg_iou)
        avg_iou_title: str = "Average IoU: " + str(avg_iou) + " =<IoU =<" + str(good_iou)
        good_iou_title: str = "Good IoU: IoU >" + str(good_iou)

        all_iou: int = sum(linkages_df.iou.notnull())

        data_bad: List[float] = [linkages_df.loc[(linkages_df.iou.notnull()) & (linkages_df.iou < avg_iou)].iou.count(),
                                 linkages_df.loc[(linkages_df.iou.notnull()) & (linkages_df.iou < avg_iou)].ecludien_distance.max(),
                                 round((linkages_df.loc[(linkages_df.iou.notnull()) & (linkages_df.iou < avg_iou)].iou.count() / all_iou)*100, 2)]

        data_avg: List[float] = [linkages_df.loc[(linkages_df.iou.notnull()) & (linkages_df.iou >= avg_iou) & (linkages_df.iou <= good_iou)].iou.count(),
                                 linkages_df.loc[(linkages_df.iou.notnull()) & (linkages_df.iou >= avg_iou) & (linkages_df.iou <= good_iou)].ecludien_distance.max(),
                                 round((linkages_df.loc[(linkages_df.iou.notnull()) & (linkages_df.iou >= avg_iou) & (linkages_df.iou <= good_iou)].iou.count(

                                 ) / all_iou)*100,2)]

        data_good: List[float] = [linkages_df.loc[(linkages_df.iou.notnull()) & (linkages_df.iou > good_iou)].iou.count(),
                                  linkages_df.loc[(linkages_df.iou.notnull()) & (linkages_df.iou > good_iou)].ecludien_distance.max(),
                                  round((linkages_df.loc[(linkages_df.iou.notnull()) & (linkages_df.iou > good_iou)].iou.count() / all_iou)*100, 2)]

        indexes: List[str] = ['Total detections', 'Maximum Euclidian Distance (in px)', 'Percentage of IoU']
        columns: List[List[float]] = [bad_iou_title, avg_iou_title, good_iou_title]

        return pd.DataFrame(data=np.array([data_bad, data_avg, data_good]).swapaxes(0, 1), columns=columns, index=indexes)

    # #   wrap text and align center and change font
    def format_table(self, sheet, max_row: int, starting_row: int, end_row: int):
        for index in range(1, max_row + 1):
            cellname = (xlsxwriter.utility.xl_col_to_name(index))
            sheet[cellname + str(starting_row)].font = Font(name="Tahoma")
            # sheet[cellname + str(starting_row)].border = Border(right=Side(style='thin'))
            for row in range(starting_row, starting_row + end_row + 2):
                sheet[cellname + str(row)].alignment = Alignment(horizontal="center", wrapText=True)
            sheet.column_dimensions[cellname].width = 30

        return sheet

    def add_border(self, sheet, starting_row, ending_row, len_column, title):

        align = Alignment(horizontal="center", wrapText=True)

        sheet.cell(row=starting_row + 1, column=1, value=title)
        sheet.cell(row=starting_row + 1, column=1).font = Font(b=True, name="Tahoma")
        sheet.cell(row=starting_row + 1, column=1).alignment = Alignment(horizontal="center", wrapText=True)

        cellname: str = (xlsxwriter.utility.xl_col_to_name(len_column))
        rows = [rows for rows in sheet["A" + str(starting_row + 1) + ":" + str(cellname) + str(ending_row)]]
        flattened = [item for sublist in rows for item in sublist]
        [(setattr(cell, 'border', self.thin_border), setattr(cell, 'alignment', align)) for cell in flattened]

        return sheet

    def _add_table_border(self, worksheet, starting_row, ending_row, column_names):
        #        add horizontal line border
        for index, column_name in enumerate(column_names):
            worksheet.cell(row=starting_row, column=index + 1).border = self.thick_border
            worksheet.cell(row=ending_row, column=index + 1).border = self.bottom_thick_border
        #       add vertical line border
        for row_index in range(starting_row + 1, ending_row):
            worksheet.cell(row=row_index, column=1).border = Border(left=Side(style='thick'))
            worksheet.cell(row=row_index, column=len(column_names)).border = Border(right=Side(style='thick'))

        #  add corner border on the last line
        worksheet.cell(row=ending_row, column=len(column_names)).border = Border(right=Side(style='thick'),
                                                                                 bottom=Side(style='thick'))

        return worksheet

    def _merge_cells(self, worksheet, linkages_df, column_names, starting_row):
        # merge cells
        all_file_name = linkages_df.image_path.unique().tolist()
        for image_name in all_file_name:
            min_index = linkages_df.loc[linkages_df.image_path == image_name].index.min()
            max_index = linkages_df.loc[linkages_df.image_path == image_name].index.max()
            start_merge_row = min_index + starting_row + 1
            end_merge_row = max_index + starting_row + 1
            worksheet.merge_cells('A' + str(start_merge_row) + ':A' + str(end_merge_row))

            for col in range(1, len(column_names) + 1):
                for i in range(start_merge_row, end_merge_row + 1):
                    worksheet.cell(row=i, column=col).border = self.thin_border
                    worksheet.cell(row=i, column=col).alignment = Alignment(horizontal="center")
                worksheet.cell(row=end_merge_row, column=col).border = self.thick_bottom

            worksheet['A' + str(start_merge_row)].alignment = Alignment(wrap_text=True, horizontal="center",
                                                                        vertical="center")

        return worksheet

    def _conditional_formating_cells(self, worksheet, column_index, starting_row, ending_row, min_color, mid_color,
                                     max_color):
        worksheet.conditional_format(str(column_index + starting_row) + ":" + str(column_index + ending_row),
                                     {'type': '2_color_scale', 'min_color': '#' + str(min_color),
                                      'mid_color': '#' + str(mid_color), 'max_color': '#' + str(max_color)})

        return worksheet

    def _format_table_content(self, worksheet, column_names, starting_row, ending_row):
        #     wrap text and align center and change font
        for index, column_name in enumerate(column_names):
            cellname = (xlsxwriter.utility.xl_col_to_name(index))
            worksheet[cellname + str(starting_row)] = str(column_name)
            worksheet[cellname + str(starting_row)].font = Font(b=True, name="Tahoma")
            worksheet[cellname + str(starting_row)].border = Border(right=Side(style='thin'))
            cellsize = len(column_name)
            worksheet.column_dimensions[cellname].width = cellsize + 10
            for row in range(starting_row, ending_row + 2):
                worksheet[cellname + str(row)].alignment = Alignment(horizontal="center", wrapText=True)
                if index == 0:
                    worksheet[cellname + str(row)].font = Font(b=True, name="Tahoma")
                if row == starting_row:
                    worksheet[cellname + str(row)].font = Font(b=True, name="Tahoma")

        return worksheet

    def create_report(self, linkages_df: DataFrame, prediction_df: DataFrame, gt_df: DataFrame, evaluation_job_parameter: EvaluationJobParameters,
                      output_path: str):
        all_columns_df = self.get_object_detection_detailed_df(linkages_df, prediction_df, gt_df)
        #  general sheet vars

        # Get needed table for General
        multi_class_confusion_df = self.get_multi_class_confusion_df(linkages_df=linkages_df)
        confusion_matrix_df = self.get_confusion_matrix_df(linkages_df=linkages_df)
        statistics_df = self.get_statistics_df(linkages_df=linkages_df, predictions_df=prediction_df, gt_df=gt_df,
                                               bad_iou=evaluation_job_parameter.iou_average_threshold)
        iou_df = self.get_iou_df(linkages_df=linkages_df, evaluation_job_parameter=evaluation_job_parameter)

        # get general sheet indexes
        general_starting_row: int = 7
        len_conf_matrix: int = len(confusion_matrix_df.index) + 1
        len_multi_class: int = len(multi_class_confusion_df.index) + 1
        len_statistics: int = len(statistics_df.index) + 1
        len_iou: int = len(iou_df.index) + 1

        # get per label index
        starting_row: int = 7
        ending_row: int = all_columns_df.index.max() + 8
        output_file_name: str = os.path.join(output_path, '5-Excel_Detailed_per_Detection.xlsx')

        column_names: List[str] = ['Image Name', 'GT label', 'Prediction Label', 'Linkage Case', 'IoU',
                                   'Euclidean Distance',
                                   'Detection confidence', 'GT centroid', 'Prediction centroid', 'GT Box left',
                                   'GT Box right',
                                   'GT Box top',
                                   'GT Box bottom', 'Prediction Box left', 'Prediction Box right', 'Prediction Box top',
                                   'Prediction Box bottom']

        max_image_column_width: int = all_columns_df.image_path.str.len().max() + 15

        writer = pd.ExcelWriter(output_file_name, engine='xlsxwriter')

        # general sheet dfs
        confusion_matrix_index: int = general_starting_row
        confusion_matrix_df.to_excel(writer, sheet_name='General', header=True, index=True, startrow=general_starting_row)

        multi_class_index: int = confusion_matrix_index + len_conf_matrix + 3
        multi_class_confusion_df.to_excel(writer, sheet_name='General', header=True, index=True, startrow=multi_class_index)

        statistics_index: int = multi_class_index + len_multi_class + 3
        statistics_df.to_excel(writer, sheet_name='General', header=True, index=True, startrow=statistics_index)

        iou_index: int = statistics_index + len_statistics + 3
        iou_df.to_excel(writer, sheet_name='General', float_format="%.2f", header=True, index=True, startrow=iou_index)

        general_ending_row: int = iou_index + len(iou_df.index) + 1

        # Per label sheet dfs
        all_columns_df.to_excel(writer, sheet_name='Per Detection Evaluation', header=False, index=False, startrow=7)

        worksheetG = writer.sheets['General']
        worksheetG.set_column(0, 0, 40)
        # worksheet.freeze_panes(0,1)

        worksheet = writer.sheets['Per Detection Evaluation']

        # # conditional formatting
        # worksheet = self._conditional_formating_cells(column_index='E', starting_row=starting_row, ending_row=ending_row,
        #                                               min_color="F20089", mid_color="A100F2", max_color="6FFFE9")
        # worksheet = self._conditional_formating_cells(column_index='F', starting_row=starting_row, ending_row=ending_row,
        #                                               min_color="F20089", mid_color="A100F2", max_color="6FFFE9")
        # worksheet = self._conditional_formating_cells(column_index='G', starting_row=starting_row, ending_row=ending_row,
        #                                               min_color="F20089", mid_color="A100F2", max_color="6FFFE9")

        # set size of 0:0 to 40
        worksheet.set_column(0, 0, 40)

        # make image_name and column_names freezed
        worksheet.freeze_panes(starting_row, 1)

        writer.save()

        wb = load_workbook(filename=output_file_name)
        worksheet = wb['Per Detection Evaluation']
        worksheetG = wb['General']

        # Worksheet GENRAL
        worksheetG = self.add_general_header_info(worksheetG, linkages_df,
                                                  evaluation_job_parameter=evaluation_job_parameter)

        max_row: int = max(len(multi_class_confusion_df.columns), len(iou_df.columns))

        # sheet=format_table(sheet,def_conf_matrix_df,starting_row,LEN_CONF_DEF)
        worksheetG = self.format_table(worksheetG, max_row, general_starting_row, general_ending_row)

        worksheetG = self.add_border(worksheetG, general_starting_row, multi_class_index - 3, len(confusion_matrix_df.columns), "Confusion Matrix:")
        worksheetG = self.add_border(worksheetG, multi_class_index, statistics_index - 3, len(multi_class_confusion_df.columns),
                                     "Multi-Class Confusion Matrix:")
        worksheetG = self.add_border(worksheetG, statistics_index, iou_index - 3, len(statistics_df.columns), "Statistics: ")
        worksheetG = self.add_border(worksheetG, iou_index, general_ending_row, len(iou_df.columns), "Evaluation of Linked Boxes:")

        # Add font and wrap text and alignment center
        worksheet = self._format_table_content(worksheet, column_names, starting_row, ending_row)

        # make image_name column larger that all text inside it
        worksheet.column_dimensions["A"].width = max_image_column_width

        # Add table border
        worksheet = self._add_table_border(worksheet, starting_row, ending_row, column_names)

        # Add Header section
        # worksheet = self._add_header_image(worksheet)
        worksheet = self._add_header_info(worksheet, evaluation_job_parameter.job_name, evaluation_job_parameter.url, evaluation_job_parameter.model_name)

        # Merge cells per image name
        worksheet = self._merge_cells(worksheet, all_columns_df, column_names, starting_row)
        largest_image_name = all_columns_df.image_path.str.len().max() + 15
        worksheet.column_dimensions["A"].width = largest_image_name

        wb.save(output_file_name)
