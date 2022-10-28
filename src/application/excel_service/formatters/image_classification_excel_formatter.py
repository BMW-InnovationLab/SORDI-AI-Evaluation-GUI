import os


from typing import List
import numpy as np
from openpyxl.styles import Alignment, Font, Border, Side
from pandas import DataFrame
from sklearn.metrics import confusion_matrix
from statistics import mean
import pandas as pd
import openpyxl
import xlsxwriter
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook

from application.excel_service.models.linkage_definition import LinkageDefinition
from domain.contracts.abstract_excel_formatter import AbstractExcelFormatter
from domain.models.evaluation_job_parameters import EvaluationJobParameters
from domain.models.linkage_type import LinkageType

class ImageClassificationExcelFormatter(AbstractExcelFormatter):

    def __init__(self):
        self.thick_border = Border(left=Side(style='thick'),
                                   right=Side(style='thick'),
                                   top=Side(style='thick'),
                                   bottom=Side(style='thin'))

        self.top_thick_border = Border(top=Side(style='thick'))
        self.bottom_thick_border = Border(bottom=Side(style='thick'))
        self.thin_border = Border(left=Side(style='thin'),
                                  right=Side(style='thin'),
                                  top=Side(style='thin'),
                                  bottom=Side(style='thin'))


        self.thick_bottom = Border(left=Side(style='thin'),
                                   right=Side(style='thin'),
                                   top=Side(style='thin'),
                                       bottom=Side(style='thick'))


    def get_image_classification_detailed_df(self,linkages_df: DataFrame):
        all_data = linkages_df[['image_path', 'gt_label', 'prediction_label', 'confidence', 'linkage_type']].copy()
        all_data = all_data[['image_path', 'gt_label', 'prediction_label', 'linkage_type', 'confidence']]
        all_data.linkage_type = all_data.linkage_type.apply(lambda x: LinkageDefinition[x].value)
        all_data.image_path = all_data.image_path.apply(lambda x: x.replace('../dataset/',''))

        all_data.sort_values('image_path', inplace=True)
        all_data.reset_index(drop=True, inplace=True)
        return all_data

    def _add_header_info(self,worksheet, evaluation_job_name:str, evaluation_url:str, evaluation_model_name:str):
        wrap_text = Alignment(wrapText=True)
        worksheet['A2'].alignment = wrap_text
        worksheet['A2'].font = Font(b=True, name="Tahoma")
        worksheet['A2'] = "Evaluation Job Per Image Report"
        worksheet['A3'] = 'Evaluation Job Name : ' + str(evaluation_job_name)
        worksheet['A4'] = 'Inference URL : ' + str(evaluation_url)
        worksheet['A5'] = 'Model Name : ' + str(evaluation_model_name)

        return worksheet


    # #   wrap text and align center and change font
    def format_table(self,sheet, max_row: int, starting_row: int, end_row: int):
        for index in range(1, max_row + 1):
            cellname = (xlsxwriter.utility.xl_col_to_name(index))
            sheet[cellname + str(starting_row)].font = Font(name="Tahoma")
            sheet[cellname + str(starting_row)].border = Border(right=Side(style='thin'))
            for row in range(starting_row, starting_row + end_row + 2):
                sheet[cellname + str(row)].alignment = Alignment(horizontal="center", wrapText=True)
            sheet.column_dimensions[cellname].width = 30

        return sheet


    def _add_table_border(self,worksheet, starting_row, ending_row, column_names):
        #        add horizontal line border
        for index, column_name in enumerate(column_names):
            worksheet.cell(row=starting_row, column=index + 1).border = self.thick_border
            worksheet.cell(row=ending_row, column=index + 1).border = self.bottom_thick_border
        #       add vertical line border
        for row_index in range(starting_row + 1, ending_row):
            worksheet.cell(row=row_index, column=1).border = Border(left=Side(style='thick'))
            worksheet.cell(row=row_index, column=len(column_names)).border = Border(right=Side(style='thick'))

        #  add corner border on the last line
        worksheet.cell(row=ending_row, column=len(column_names)).border = Border(right=Side(style='thick'),
                                                                                 bottom=Side(style='thick'))

        return worksheet


    def _format_table_content(self,worksheet, column_names:List[str], starting_row:int, ending_row:int):
        #     wrap text and align center and change font
        for index, column_name in enumerate(column_names):
            cellname = (xlsxwriter.utility.xl_col_to_name(index))
            worksheet[cellname + str(starting_row)] = str(column_name)
            worksheet[cellname + str(starting_row)].font = Font(b=True, name="Tahoma")
            worksheet[cellname + str(starting_row)].border = Border(right=Side(style='thin'))
            cellsize = len(column_name)
            worksheet.column_dimensions[cellname].width = cellsize + 10
            for row in range(starting_row, ending_row + 2):
                worksheet[cellname + str(row)].alignment = Alignment(horizontal="center", wrapText=True)
                if index == 0:
                    worksheet[cellname + str(row)].font = Font(b=True, name="Tahoma")
                if row == starting_row:
                    worksheet[cellname + str(row)].font = Font(b=True, name="Tahoma")

        return worksheet

    def add_general_header_info(self, worksheet, linkages_df: DataFrame, evaluation_job_parameter: EvaluationJobParameters):
        wrap_text = Alignment(wrapText=True)
        worksheet['A1'].alignment = wrap_text
        worksheet['A1'].font = Font(b=True, name="Tahoma")
        worksheet['A1'] = "Evaluation Job Per Image Report"
        worksheet['A2'] = 'Evaluation Job Name : ' + str(evaluation_job_parameter.job_name)
        worksheet['A3'] = 'Inference URL : ' + str(evaluation_job_parameter.url)
        worksheet['A4'] = 'Model Name : ' + str(evaluation_job_parameter.model_name)
        worksheet['A5'] = 'Total Image Tested : ' + str(len(linkages_df.image_path.unique()))
        for i in range(1, 6):
            worksheet.cell(row=i, column=1).border =self.thin_border

        return worksheet

    def get_multi_class_confusion_df(self,linkages_df: DataFrame) -> DataFrame:
        prediction_labels: List[str] = linkages_df.prediction_label.astype(str).tolist()
        gt_labels: List[str] = linkages_df.gt_label.astype(str).tolist()
        all_labels: List[str] = list(set(prediction_labels + gt_labels))

        all_labels.sort()
        confusion_matrix_output = confusion_matrix(gt_labels, prediction_labels, labels=all_labels).tolist()
        return pd.DataFrame(data=confusion_matrix_output, columns=all_labels, index=all_labels)

    def get_confusion_matrix_df(self,linkages_df: DataFrame) -> DataFrame:
        columns: List[str] = ['No Prediction', 'Prediction']
        indexes: List[str] = ['No True Label', 'True Label']
        labels: List[str] = [linkage_type.value for linkage_type in LinkageDefinition]
        data = [(linkages_df.linkage_type.values == label).sum().tolist() for label in labels]
        data = np.asarray(data).reshape(2, 2)[::-1]
        data = np.fliplr(data)
        return pd.DataFrame(data=data, columns=columns, index=indexes)

    def get_statistics_df(self,linkages_df: DataFrame, predictions_df: DataFrame, gt_df: DataFrame) -> DataFrame:
        TP, TN, FP, FN = [(linkages_df.linkage_type.values == case.value).sum() for case in LinkageDefinition]

        try:
            precision: float = round(TP / (TP + FP), 2)
            recall: float = round(TP / (TP + FN), 2)
            f_score: float = round(2 * precision * recall / (precision + recall), 2)
        except Exception:
            precision = recall = f_score = 0

        try:
            accuracy = round((TP) / float(TP  + FP + FN), 2)
        except Exception:
            accuracy = 0
        try:
            top_acc = round(linkages_df.confidence.max(), 2)
            avg_conf = round(mean(linkages_df.loc[linkages_df.confidence.notnull(), 'confidence'].tolist()), 2)
        except Exception:
            top_acc =avg_conf =0

        len_gt = len(gt_df.index)
        len_pred = len(predictions_df.index)

        indexes = ['Top-1 Accuaracy', 'Total number of predictions', 'Total number of Ground truths',
                   'Accuracy', 'Precision', 'Recall', 'F-score',
                   'Average Confidence']
        data = [top_acc, len_gt, len_pred, accuracy, precision, recall, f_score, avg_conf]
        return pd.DataFrame(data=data, index=indexes, columns=['score'])

    # #   wrap text and  center and change font
    def format_table(self,sheet, max_row: int, starting_row: int, end_row: int):
        for index in range(1, max_row + 1):
            cellname = (xlsxwriter.utility.xl_col_to_name(index))
            sheet[cellname + str(starting_row)].font = Font(name="Tahoma")
            for row in range(starting_row, starting_row + end_row + 2):
                sheet[cellname + str(row)].alignment = Alignment(horizontal="center", wrapText=True)
            sheet.column_dimensions[cellname].width = 30

        return sheet

    def add_border(self,sheet, starting_row:int, ending_row:int, len_column:int, title:str):



        align = Alignment(horizontal="center", wrapText=True)

        sheet.cell(row=starting_row + 1, column=1, value=title)
        sheet.cell(row=starting_row + 1, column=1).font = Font(b=True, name="Tahoma")
        sheet.cell(row=starting_row + 1, column=1).alignment = Alignment(horizontal="center", wrapText=True)

        cellname: str = (xlsxwriter.utility.xl_col_to_name(len_column))
        rows = [rows for rows in sheet["A" + str(starting_row + 1) + ":" + str(cellname) + str(ending_row)]]
        flattened = [item for sublist in rows for item in sublist]
        [(setattr(cell, 'border', self.thin_border), setattr(cell, 'alignment', align)) for cell in flattened]

        return sheet


    def create_report(self,linkages_df:DataFrame, prediction_df:DataFrame, gt_df:DataFrame, evaluation_job_parameter: EvaluationJobParameters, output_path:str):

        linkages_df = self.get_image_classification_detailed_df(linkages_df).copy()

        # general sheet
        multi_class_confusion_df = self.get_multi_class_confusion_df(linkages_df=linkages_df)
        confusion_matrix_df = self.get_confusion_matrix_df(linkages_df=linkages_df)
        statistics_df = self.get_statistics_df(linkages_df=linkages_df, predictions_df=prediction_df, gt_df=gt_df)

        # get general sheet indexes
        general_starting_row: int = 7
        len_conf_matrix: int = len(confusion_matrix_df.index) + 1
        len_multi_class: int = len(multi_class_confusion_df.index) + 1
        len_statistics: int = len(statistics_df.index) + 1


        output_file_name: str = os.path.join(output_path, '3-Excel_Detailed_per_Classification.xlsx')
        # per label section
        # get per label index
        starting_row: int = 7
        ending_row: int = linkages_df.index.max() + 8

        column_names = ['Image Name', 'GT label', 'Prediction Label', 'Linkage Case', 'Detection confidence', ]
        max_image_column_width = linkages_df.image_path.str.len().max() + 15

        writer = pd.ExcelWriter(output_file_name, engine='xlsxwriter')

        # general sheet dfs
        confusion_matrix_index: int = general_starting_row
        confusion_matrix_df.to_excel(writer, sheet_name='General', header=True, index=True,startrow=general_starting_row)
        multi_class_index: int = confusion_matrix_index + len_conf_matrix + 3
        multi_class_confusion_df.to_excel(writer, sheet_name='General', header=True, index=True,startrow=multi_class_index)
        statistics_index: int = multi_class_index + len_multi_class + 3
        statistics_df.to_excel(writer, sheet_name='General', header=True, index=True, startrow=statistics_index)
        general_ending_row: int = statistics_index + len_statistics + 3

        # Per label sheet dfs
        linkages_df.to_excel(writer, sheet_name='Per classification Evaluation', header=False, float_format="%.2f",index=False, startrow=7)

        worksheet = writer.sheets['Per classification Evaluation']
        # set size of 0:0 to 40
        worksheet.set_column(0, 0, 40)
        # make image_name and column_names freezed
        worksheet.freeze_panes(starting_row, 1)

        worksheetG = writer.sheets['General']

        worksheetG.set_column(0, 0, 40)

        writer.save()

        wb = load_workbook(filename=output_file_name)
        worksheet = wb['Per classification Evaluation']
        worksheetG = wb['General']

        # Add font and wrap text and alignment center
        worksheet = self._format_table_content(worksheet, column_names, starting_row, ending_row)
        # make image_name column larger that all text inside it
        worksheet.column_dimensions["A"].width = max_image_column_width

        # Add table border
        worksheet = self._add_table_border(worksheet, starting_row, ending_row, column_names)
        # Add Header section
        worksheet = self._add_header_info(worksheet,  evaluation_job_parameter.job_name, evaluation_job_parameter.url,evaluation_job_parameter.model_name)

        # Merge cells per image name
        largest_image_name = linkages_df.image_path.str.len().max() + 15
        worksheet.column_dimensions["A"].width = largest_image_name


        # Worksheet GENERAL
        worksheetG = self.add_general_header_info(worksheetG, linkages_df,evaluation_job_parameter)

        max_row: int = len(multi_class_confusion_df.columns)

        worksheetG = self.format_table(worksheetG, max_row, general_starting_row, general_ending_row)

        worksheetG = self.add_border(worksheetG, general_starting_row, multi_class_index - 3,len(confusion_matrix_df.columns), "Confusion Matrix:")
        worksheetG = self.add_border(worksheetG, multi_class_index, statistics_index - 3,len(multi_class_confusion_df.columns), "Multi-Class Confusion Matrix:")
        worksheetG = self.add_border(worksheetG, statistics_index, general_ending_row - 3, len(statistics_df.columns),"Statistics: ")

        wb.save(output_file_name)


